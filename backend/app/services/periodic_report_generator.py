"""
Periodic Report Generator (Monthly / Quarterly).

Produces a ZIP archive containing:
  - report-<label>.html  : self-contained dark-theme HTML with SVG charts
  - report-<label>.xlsx  : Excel workbook with 7 sheets

7 sections:
  1. License Management
  2. User Usage
  3. Copilot Usage Statistics
  4. Technical Metrics (Model / IDE / Language / Feature)
  5. By-Org / Team Statistics
  6. Usage Trends (DAU / WAU / MAU)
  7. License Optimization (inactive / low-usage recommendations)
"""
from __future__ import annotations

import io
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from html import escape
from typing import Any

# ── palette (matches existing report_generator.py) ────────────────────────────
_ACCENT = "#539bf5"
_GREEN  = "#3fb950"
_YELLOW = "#d29922"
_RED    = "#e5534b"
_MUTED  = "#768390"
_BG     = "#0d1117"
_CARD   = "#161b22"
_CARD2  = "#21262d"
_BORDER = "#30363d"
_TEXT   = "#e6edf3"


# ── small helpers ──────────────────────────────────────────────────────────────

def _e(v: Any) -> str:
    return escape(str(v) if v is not None else "")


def _num(v: Any) -> str:
    try:
        return f"{int(float(v)):,}"
    except (ValueError, TypeError):
        return str(v)


def _pct(v: Any) -> str:
    try:
        return f"{float(v):.1f}%"
    except (ValueError, TypeError):
        return "-"


def _money(v: Any) -> str:
    try:
        return f"${float(v):,.2f}"
    except (ValueError, TypeError):
        return "-"


# ── period helpers ─────────────────────────────────────────────────────────────

def _period_months(period_type: str, year: int, period: int) -> list[str]:
    """Return list of 'YYYY-MM' strings for the given period."""
    if period_type == "quarterly":
        start_m = (period - 1) * 3 + 1
        return [f"{year:04d}-{m:02d}" for m in range(start_m, start_m + 3)]
    return [f"{year:04d}-{period:02d}"]


def _period_label(period_type: str, year: int, period: int) -> str:
    if period_type == "quarterly":
        return f"{year}-Q{period}"
    import calendar
    return f"{year}-{calendar.month_abbr[period]}"


def _in_period(day: str, months: list[str]) -> bool:
    """Check if an ISO date string 'YYYY-MM-DD' falls within the period months."""
    if not day or len(day) < 7:
        return False
    return day[:7] in months


def _in_period_ts(ts: str | None, months: list[str]) -> bool:
    """Check if an ISO timestamp falls within period months."""
    if not ts:
        return False
    try:
        d = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        return f"{d.year:04d}-{d.month:02d}" in months
    except (ValueError, TypeError):
        return False


# ── SVG line chart ─────────────────────────────────────────────────────────────

def _svg_line(data: list[dict], x_key: str, y_key: str, y_label: str,
              chart_id: str, w: int = 760, h: int = 180) -> str:
    if not data:
        return '<p class="no-data">No data.</p>'
    values = [float(d.get(y_key, 0) or 0) for d in data]
    labels = [str(d.get(x_key, "")) for d in data]
    n = len(values)
    PL, PR, PT, PB = 58, 12, 12, 40
    iw, ih = w - PL - PR, h - PT - PB
    max_v = max(values) if values else 1
    min_v = min(values) if values else 0
    rng = (max_v - min_v) or 1

    def xp(i): return PL + (i / max(n - 1, 1)) * iw
    def yp(v): return PT + ih - ((v - min_v) / rng) * ih

    pts = " ".join(f"{xp(i):.1f},{yp(v):.1f}" for i, v in enumerate(values))
    area = f"{xp(0):.1f},{PT+ih:.1f} {pts} {xp(n-1):.1f},{PT+ih:.1f}"

    n_ticks = 4
    grid = ""
    for ti in range(n_ticks + 1):
        v = min_v + rng * ti / n_ticks
        y = yp(v)
        lbl = f"{v/1000:.1f}k" if abs(v) >= 1000 else f"{v:.0f}"
        grid += (f'<line x1="{PL}" y1="{y:.1f}" x2="{w-PR}" y2="{y:.1f}" '
                 f'stroke="{_BORDER}" stroke-dasharray="4,3"/>'
                 f'<text x="{PL-4}" y="{y+4:.1f}" text-anchor="end" '
                 f'font-size="10" fill="{_MUTED}">{_e(lbl)}</text>')

    step = max(1, n // 8)
    xlbls = ""
    for i, lbl in enumerate(labels):
        if i % step == 0 or i == n - 1:
            xlbls += (f'<text x="{xp(i):.1f}" y="{PT+ih+16}" text-anchor="middle" '
                      f'font-size="10" fill="{_MUTED}">{_e(lbl[:10])}</text>')

    dots = ""
    if n <= 60:
        dots = "".join(f'<circle cx="{xp(i):.1f}" cy="{yp(v):.1f}" r="2.5" fill="{_ACCENT}"/>'
                       for i, v in enumerate(values))

    gid = f"g{chart_id}"
    return (
        f'<svg viewBox="0 0 {w} {h}" xmlns="http://www.w3.org/2000/svg" '
        f'style="width:100%;height:{h}px;display:block;background:{_CARD2};border-radius:6px">'
        f'<defs><linearGradient id="{gid}" x1="0" y1="0" x2="0" y2="1">'
        f'<stop offset="0%" stop-color="{_ACCENT}" stop-opacity="0.25"/>'
        f'<stop offset="100%" stop-color="{_ACCENT}" stop-opacity="0.02"/>'
        f'</linearGradient></defs>'
        f'{grid}{xlbls}'
        f'<polygon points="{area}" fill="url(#{gid})"/>'
        f'<polyline points="{pts}" fill="none" stroke="{_ACCENT}" stroke-width="2" '
        f'stroke-linejoin="round" stroke-linecap="round"/>'
        f'{dots}'
        f'<text x="{PL+iw/2:.0f}" y="{h-4}" text-anchor="middle" '
        f'font-size="10" fill="{_MUTED}">{_e(y_label)}</text>'
        f'</svg>'
    )


# ── KPI card ───────────────────────────────────────────────────────────────────

def _kpi(label: str, value: str, color: str = _ACCENT, sub: str = "") -> str:
    sub_html = f'<div style="font-size:11px;color:{_MUTED};margin-top:2px">{_e(sub)}</div>' if sub else ""
    return (f'<div style="background:{_CARD2};border:1px solid {_BORDER};border-radius:8px;'
            f'padding:14px 18px;min-width:120px;flex:1">'
            f'<div style="font-size:22px;font-weight:700;color:{color}">{_e(value)}</div>'
            f'<div style="font-size:12px;color:{_MUTED};margin-top:4px">{_e(label)}</div>'
            f'{sub_html}</div>')


def _kpi_row(*cards: str) -> str:
    return f'<div style="display:flex;gap:12px;flex-wrap:wrap;margin:12px 0">{"".join(cards)}</div>'


# ── HTML table helper ──────────────────────────────────────────────────────────

def _table(headers: list[str], rows: list[list[Any]], max_rows: int = 500) -> str:
    th = "".join(
        f'<th style="padding:8px 12px;text-align:left;color:{_MUTED};'
        f'font-size:12px;font-weight:600;border-bottom:1px solid {_BORDER}">{_e(h)}</th>'
        for h in headers
    )
    tr_list = []
    for i, row in enumerate(rows[:max_rows]):
        bg = _CARD if i % 2 == 0 else _CARD2
        tds = "".join(
            f'<td style="padding:7px 12px;font-size:13px;border-bottom:1px solid {_BORDER}">'
            f'{_e(c)}</td>'
            for c in row
        )
        tr_list.append(f'<tr style="background:{bg}">{tds}</tr>')
    note = ""
    if len(rows) > max_rows:
        note = f'<p style="color:{_MUTED};font-size:12px;margin:8px 0">... {len(rows)-max_rows} more rows (see Excel for full data)</p>'
    return (
        f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse">'
        f'<thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(tr_list)}</tbody>'
        f'</table></div>{note}'
    )


# ── section wrapper ────────────────────────────────────────────────────────────

def _section(title: str, content: str, icon: str = "", anchor: str = "") -> str:
    anchor_attr = f' id="{anchor}"' if anchor else ""
    return (
        f'<details open{anchor_attr} style="margin:20px 0">'
        f'<summary style="cursor:pointer;font-size:16px;font-weight:700;color:{_TEXT};'
        f'padding:10px 0;border-bottom:2px solid {_ACCENT}">{icon} {_e(title)}</summary>'
        f'<div style="padding:16px 0">{content}</div>'
        f'</details>'
    )


# ── CSS ────────────────────────────────────────────────────────────────────────

_CSS = f"""
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
       background: {_BG}; color: {_TEXT}; padding: 24px; }}
h1 {{ font-size: 24px; color: {_ACCENT}; margin-bottom: 4px; }}
.subtitle {{ color: {_MUTED}; font-size: 13px; margin-bottom: 24px; }}
.no-data {{ color: {_MUTED}; font-style: italic; font-size: 13px; padding: 12px 0; }}
a {{ color: {_ACCENT}; text-decoration: none; }}
"""


# ══════════════════════════════════════════════════════════════════════════════
# DATA AGGREGATION
# ══════════════════════════════════════════════════════════════════════════════

def _aggregate(
    data_collector,
    period_type: str,
    year: int,
    period: int,
    org_filter: list[str],
    all_scope_names: list[str] | None = None,
) -> dict:
    """Collect and aggregate all data needed for the report."""
    months = _period_months(period_type, year, period)
    label = _period_label(period_type, year, period)
    now = datetime.now(timezone.utc)

    # ── determine which orgs to include ───────────────────────────────────────
    if all_scope_names is None:
        # Fallback: try importing api_manager (works when running inside the app)
        try:
            from ..services.api_manager import api_manager as _am  # type: ignore
            all_scope_names = []
            for o in _am.get_all_orgs():
                all_scope_names.append(o["login"])
            for ent in _am.get_all_enterprises():
                slug = ent.get("slug", "")
                if slug and slug not in all_scope_names:
                    all_scope_names.append(slug)
        except ImportError:
            all_scope_names = []

    scopes = [s for s in all_scope_names if not org_filter or s in org_filter]

    # ── Section 1: License Management ─────────────────────────────────────────
    license_rows: list[dict] = []   # per-org
    total_seats_all = 0
    total_active_all = 0
    total_cost_all = 0.0

    # team breakdown: {org: {team: {total, active}}}
    team_map: dict[str, dict[str, dict]] = defaultdict(lambda: defaultdict(lambda: {"total": 0, "active": 0}))

    for scope in scopes:
        billing = data_collector.load_latest("billing", scope)
        seats_data = data_collector.load_latest("seats", scope)
        if not billing and not seats_data:
            continue

        if billing:
            price = billing.get("_detected_price_per_seat", 19.0)
            sb = billing.get("seat_breakdown", {})
            total_s = sb.get("total", 0)
            active_s = sb.get("active_this_cycle", 0)
            plan = billing.get("_detected_plan_type", "unknown")
        else:
            price = 39.0
            plan = "unknown"
            total_s = seats_data.get("total_seats", 0) if seats_data else 0
            active_s = 0

        # Count active from seats data filtered by period
        if seats_data:
            for s in seats_data.get("seats", []):
                last = s.get("last_activity_at")
                assignee = (s.get("assignee") or {}).get("login", "")
                teams_list = s.get("assigning_team")
                team_name = (teams_list.get("name", "") if isinstance(teams_list, dict)
                             else (teams_list[0].get("name", "") if isinstance(teams_list, list) and teams_list else ""))

                is_active_in_period = _in_period_ts(last, months)
                team_map[scope][team_name or "(no team)"]["total"] += 1
                if is_active_in_period:
                    team_map[scope][team_name or "(no team)"]["active"] += 1

        total_seats_all += total_s
        total_active_all += active_s
        total_cost_all += total_s * price

        license_rows.append({
            "org": scope, "plan": plan, "price_per_seat": price,
            "total_seats": total_s, "active_seats": active_s,
            "inactive_seats": total_s - active_s,
            "utilization_pct": round(active_s / total_s * 100, 1) if total_s > 0 else 0.0,
            "monthly_cost": round(total_s * price, 2),
        })

    # flat team rows
    team_rows: list[dict] = []
    for scope, teams in team_map.items():
        for team, td in teams.items():
            team_rows.append({
                "org": scope, "team": team,
                "total": td["total"], "active": td["active"],
                "inactive": td["total"] - td["active"],
                "utilization_pct": round(td["active"] / td["total"] * 100, 1) if td["total"] > 0 else 0.0,
            })

    # ── Section 2: User Usage ──────────────────────────────────────────────────
    user_usage: list[dict] = []
    for scope in scopes:
        seats_data = data_collector.load_latest("seats", scope)
        uu_data = data_collector.load_latest("usage_users", scope)

        # build per-user usage from usage_users
        user_day_agg: dict[str, dict] = defaultdict(lambda: {
            "days_active": 0, "interactions": 0, "code_gen": 0, "code_accept": 0,
            "weeks": set(), "monthly_set": set(),
        })
        if uu_data:
            for rec in uu_data.get("records", []):
                day = rec.get("day", "")
                if not _in_period(day, months):
                    continue
                login = rec.get("user_login", "")
                if not login:
                    continue
                u = user_day_agg[login]
                u["days_active"] += 1
                u["interactions"] += rec.get("user_initiated_interaction_count", 0)
                u["code_gen"] += rec.get("code_generation_activity_count", 0)
                u["code_accept"] += rec.get("code_acceptance_activity_count", 0)
                if day:
                    try:
                        dt = datetime.strptime(day, "%Y-%m-%d")
                        u["weeks"].add(dt.isocalendar()[1])
                        u["monthly_set"].add(day[:7])
                    except ValueError:
                        pass

        if seats_data:
            for s in seats_data.get("seats", []):
                assignee = (s.get("assignee") or {})
                login = assignee.get("login", "")
                if not login:
                    continue
                last = s.get("last_activity_at")
                created = s.get("created_at", "")
                editor = s.get("last_activity_editor", "")
                teams_obj = s.get("assigning_team")
                team_name = (teams_obj.get("name", "") if isinstance(teams_obj, dict)
                             else (teams_obj[0].get("name", "") if isinstance(teams_obj, list) and teams_obj else ""))

                u_agg = user_day_agg.get(login, {})
                days_active = u_agg.get("days_active", 0)
                weeks_active = len(u_agg.get("weeks", set()))
                months_active = len(u_agg.get("monthly_set", set()))
                interactions = u_agg.get("interactions", 0)
                code_gen = u_agg.get("code_gen", 0)
                code_accept = u_agg.get("code_accept", 0)

                user_usage.append({
                    "login": login,
                    "org": scope,
                    "team": team_name,
                    "created_at": created[:10] if created else "",
                    "last_active": last[:10] if last else "",
                    "last_editor": editor,
                    "days_active": days_active,
                    "weeks_active": weeks_active,
                    "months_active": months_active,
                    "interactions": interactions,
                    "code_suggestions": code_gen,
                    "accepted_suggestions": code_accept,
                    "acceptance_rate": round(code_accept / code_gen * 100, 1) if code_gen > 0 else 0.0,
                    "avg_interactions_per_active_day": round(interactions / days_active, 1) if days_active > 0 else 0.0,
                    "usage_frequency": (
                        "Cao" if days_active >= 15
                        else "Trung bình" if days_active >= 5
                        else "Thấp" if days_active > 0
                        else "Không dùng"
                    ),
                })

    user_usage.sort(key=lambda x: -(x["days_active"] + x["interactions"]))

    # ── Section 3: Copilot Usage Statistics ───────────────────────────────────
    total_interactions = 0
    total_code_gen = 0
    total_code_accept = 0
    total_loc_suggested = 0
    total_loc_accepted = 0
    total_chat_interactions = 0
    total_chat_users_sum = 0
    total_pr_summary_interactions = 0
    total_dau_sum = 0
    daily_active_days = 0
    _CHAT_FEATURES = {"chat", "chat_panel", "chat_panel_agent_mode", "chat_panel_unknown_mode",
                      "copilot_chat", "inline_chat", "quick_chat"}

    for scope in scopes:
        usage = data_collector.load_latest("usage", scope)
        if not usage:
            continue
        for rec in usage.get("records", []):
            for dt in rec.get("day_totals", []):
                day = dt.get("day", "")
                if not _in_period(day, months):
                    continue
                total_interactions += dt.get("user_initiated_interaction_count", 0)
                total_code_gen += dt.get("code_generation_activity_count", 0)
                total_code_accept += dt.get("code_acceptance_activity_count", 0)
                total_loc_suggested += (dt.get("loc_suggested_to_add_sum", 0) +
                                        dt.get("loc_suggested_to_delete_sum", 0))
                total_loc_accepted += (dt.get("loc_added_sum", 0) +
                                       dt.get("loc_deleted_sum", 0))
                total_chat_users_sum += dt.get("monthly_active_chat_users", 0)
                total_dau_sum += dt.get("daily_active_users", 0)
                daily_active_days += 1
                # extract chat & PR summary from feature breakdown
                for fb in dt.get("totals_by_feature", []):
                    fname = (fb.get("feature") or "").lower()
                    if fname in _CHAT_FEATURES or "chat" in fname:
                        total_chat_interactions += fb.get("user_initiated_interaction_count", 0)
                    if "pr_summary" in fname or "pull_request" in fname:
                        total_pr_summary_interactions += fb.get("user_initiated_interaction_count", 0)

    avg_dau = round(total_dau_sum / daily_active_days, 1) if daily_active_days > 0 else 0
    acceptance_rate = round(total_code_accept / total_code_gen * 100, 1) if total_code_gen > 0 else 0.0
    # Peak monthly_active_chat_users (take max day value as representative)
    peak_chat_users = 0
    for scope in scopes:
        usage = data_collector.load_latest("usage", scope)
        if not usage:
            continue
        for rec in usage.get("records", []):
            for dt in rec.get("day_totals", []):
                if _in_period(dt.get("day", ""), months):
                    peak_chat_users = max(peak_chat_users, dt.get("monthly_active_chat_users", 0))

    # ── Section 4: Technical Metrics ──────────────────────────────────────────
    model_map: dict[str, dict] = defaultdict(lambda: {"interactions": 0, "code_gen": 0, "code_accept": 0})
    ide_map: dict[str, dict] = defaultdict(lambda: {"interactions": 0, "code_gen": 0, "code_accept": 0})
    lang_map: dict[str, dict] = defaultdict(lambda: {"code_gen": 0, "code_accept": 0, "loc_suggested": 0, "loc_accepted": 0})
    feature_map: dict[str, dict] = defaultdict(lambda: {"interactions": 0, "code_gen": 0, "code_accept": 0})

    for scope in scopes:
        usage = data_collector.load_latest("usage", scope)
        if not usage:
            continue
        for rec in usage.get("records", []):
            for dt in rec.get("day_totals", []):
                if not _in_period(dt.get("day", ""), months):
                    continue
                for mb in dt.get("totals_by_model_feature", []):
                    m = mb.get("model", "unknown")
                    model_map[m]["interactions"] += mb.get("user_initiated_interaction_count", 0)
                    model_map[m]["code_gen"] += mb.get("code_generation_activity_count", 0)
                    model_map[m]["code_accept"] += mb.get("code_acceptance_activity_count", 0)
                for ib in dt.get("totals_by_ide", []):
                    ide = ib.get("ide", "unknown")
                    ide_map[ide]["interactions"] += ib.get("user_initiated_interaction_count", 0)
                    ide_map[ide]["code_gen"] += ib.get("code_generation_activity_count", 0)
                    ide_map[ide]["code_accept"] += ib.get("code_acceptance_activity_count", 0)
                for lb in dt.get("totals_by_language_feature", []):
                    lang = lb.get("language", "unknown")
                    lang_map[lang]["code_gen"] += lb.get("code_generation_activity_count", 0)
                    lang_map[lang]["code_accept"] += lb.get("code_acceptance_activity_count", 0)
                    lang_map[lang]["loc_suggested"] += lb.get("loc_suggested_to_add_sum", 0) + lb.get("loc_suggested_to_delete_sum", 0)
                    lang_map[lang]["loc_accepted"] += lb.get("loc_added_sum", 0) + lb.get("loc_deleted_sum", 0)
                for fb in dt.get("totals_by_feature", []):
                    f = fb.get("feature", "unknown")
                    feature_map[f]["interactions"] += fb.get("user_initiated_interaction_count", 0)
                    feature_map[f]["code_gen"] += fb.get("code_generation_activity_count", 0)
                    feature_map[f]["code_accept"] += fb.get("code_acceptance_activity_count", 0)

    model_rows = sorted(
        [{"model": k, **v, "acceptance_rate": round(v["code_accept"]/v["code_gen"]*100,1) if v["code_gen"]>0 else 0}
         for k, v in model_map.items()],
        key=lambda x: -x["interactions"]
    )
    ide_rows = sorted(
        [{"ide": k, **v, "acceptance_rate": round(v["code_accept"]/v["code_gen"]*100,1) if v["code_gen"]>0 else 0}
         for k, v in ide_map.items()],
        key=lambda x: -x["interactions"]
    )
    lang_rows = sorted(
        [{"language": k, **v, "acceptance_rate": round(v["code_accept"]/v["code_gen"]*100,1) if v["code_gen"]>0 else 0}
         for k, v in lang_map.items()],
        key=lambda x: -x["code_gen"]
    )
    feature_rows = sorted(
        [{"feature": k, **v, "acceptance_rate": round(v["code_accept"]/v["code_gen"]*100,1) if v["code_gen"]>0 else 0}
         for k, v in feature_map.items()],
        key=lambda x: -x["interactions"]
    )

    # ── Section 5: By-Org / Team Statistics ───────────────────────────────────
    org_stats: list[dict] = []
    for scope in scopes:
        usage = data_collector.load_latest("usage", scope)
        uu_data = data_collector.load_latest("usage_users", scope)
        billing = data_collector.load_latest("billing", scope)
        seats_data = data_collector.load_latest("seats", scope)

        active_users_period: set[str] = set()
        interactions_scope = 0
        code_gen_scope = 0
        code_accept_scope = 0

        if usage:
            for rec in usage.get("records", []):
                for dt in rec.get("day_totals", []):
                    if not _in_period(dt.get("day", ""), months):
                        continue
                    interactions_scope += dt.get("user_initiated_interaction_count", 0)
                    code_gen_scope += dt.get("code_generation_activity_count", 0)
                    code_accept_scope += dt.get("code_acceptance_activity_count", 0)

        if uu_data:
            for rec in uu_data.get("records", []):
                if _in_period(rec.get("day", ""), months):
                    login = rec.get("user_login", "")
                    if login:
                        active_users_period.add(login)

        if billing:
            price = billing.get("_detected_price_per_seat", 19.0)
            sb = billing.get("seat_breakdown", {})
            total_s = sb.get("total", 0)
        elif seats_data:
            price = 39.0
            total_s = seats_data.get("total_seats", 0)
        else:
            continue

        org_stats.append({
            "org": scope,
            "total_seats": total_s,
            "active_users_period": len(active_users_period),
            "utilization_pct": round(len(active_users_period) / total_s * 100, 1) if total_s > 0 else 0.0,
            "interactions": interactions_scope,
            "code_suggestions": code_gen_scope,
            "accepted_suggestions": code_accept_scope,
            "acceptance_rate": round(code_accept_scope / code_gen_scope * 100, 1) if code_gen_scope > 0 else 0.0,
            "monthly_cost": round(total_s * price, 2),
        })
    org_stats.sort(key=lambda x: -x["active_users_period"])
    # Add rank after sorting
    for i, row in enumerate(org_stats, 1):
        row["rank"] = i

    # ── Team-level usage stats (cross-ref user_usage + seat team assignment) ──
    # Build: (org, team) -> {active_users: set, interactions, code_gen, code_accept, total_seats}
    team_usage_map: dict[tuple, dict] = defaultdict(lambda: {
        "active_users": set(), "interactions": 0, "code_gen": 0, "code_accept": 0, "total_seats": 0
    })
    for u in user_usage:
        key = (u["org"], u["team"] or "(no team)")
        t = team_usage_map[key]
        if u["days_active"] > 0:
            t["active_users"].add(u["login"])
        t["interactions"] += u["interactions"]
        t["code_gen"] += u["code_suggestions"]
        t["code_accept"] += u["accepted_suggestions"]
    # Merge seat counts from team_rows already computed
    for tr in team_rows:
        key = (tr["org"], tr["team"])
        team_usage_map[key]["total_seats"] = tr["total"]

    team_usage_rows = sorted(
        [{
            "org": k[0], "team": k[1],
            "total_seats": v["total_seats"],
            "active_users": len(v["active_users"]),
            "utilization_pct": round(len(v["active_users"]) / v["total_seats"] * 100, 1)
                               if v["total_seats"] > 0 else 0.0,
            "interactions": v["interactions"],
            "code_suggestions": v["code_gen"],
            "accepted_suggestions": v["code_accept"],
            "acceptance_rate": round(v["code_accept"] / v["code_gen"] * 100, 1)
                               if v["code_gen"] > 0 else 0.0,
        } for k, v in team_usage_map.items()],
        key=lambda x: -(x["interactions"] + x["active_users"])
    )

    # ── Section 6: Usage Trends ────────────────────────────────────────────────
    trend_map: dict[str, dict] = {}
    for scope in scopes:
        usage = data_collector.load_latest("usage", scope)
        if not usage:
            continue
        for rec in usage.get("records", []):
            for dt in rec.get("day_totals", []):
                day = dt.get("day", "")
                if not _in_period(day, months):
                    continue
                if day not in trend_map:
                    trend_map[day] = {"day": day, "dau": 0, "wau": 0, "mau": 0,
                                      "interactions": 0, "code_gen": 0}
                d = trend_map[day]
                d["dau"] += dt.get("daily_active_users", 0)
                d["wau"] += dt.get("weekly_active_users", 0)
                d["mau"] += dt.get("monthly_active_users", 0)
                d["interactions"] += dt.get("user_initiated_interaction_count", 0)
                d["code_gen"] += dt.get("code_generation_activity_count", 0)

    daily_trend = sorted(trend_map.values(), key=lambda x: x["day"])

    # Compute week-over-week growth rate for DAU and interactions
    for i, row in enumerate(daily_trend):
        if i >= 7:
            prev_dau = daily_trend[i - 7]["dau"]
            prev_int = daily_trend[i - 7]["interactions"]
            row["dau_wow_pct"] = round((row["dau"] - prev_dau) / prev_dau * 100, 1) if prev_dau > 0 else None
            row["int_wow_pct"] = round((row["interactions"] - prev_int) / prev_int * 100, 1) if prev_int > 0 else None
        else:
            row["dau_wow_pct"] = None
            row["int_wow_pct"] = None

    # Aggregate weekly summary for cleaner trend view
    weekly_map: dict[str, dict] = defaultdict(lambda: {"week": "", "dau_avg": 0, "interactions": 0, "code_gen": 0, "days": 0})
    for row in daily_trend:
        try:
            dt_obj = datetime.strptime(row["day"], "%Y-%m-%d")
            iso = dt_obj.isocalendar()
            week_key = f"{iso[0]}-W{iso[1]:02d}"
        except ValueError:
            week_key = row["day"][:7]
        w = weekly_map[week_key]
        w["week"] = week_key
        w["dau_avg"] += row["dau"]
        w["interactions"] += row["interactions"]
        w["code_gen"] += row["code_gen"]
        w["days"] += 1
    weekly_trend = []
    prev_int = None
    for wk, v in sorted(weekly_map.items()):
        avg_dau = round(v["dau_avg"] / v["days"], 1) if v["days"] > 0 else 0
        growth = None
        if prev_int is not None and prev_int > 0:
            growth = round((v["interactions"] - prev_int) / prev_int * 100, 1)
        weekly_trend.append({
            "week": v["week"],
            "avg_dau": avg_dau,
            "interactions": v["interactions"],
            "code_gen": v["code_gen"],
            "growth_pct": growth,
        })
        prev_int = v["interactions"]

    # ── Section 7: License Optimization ───────────────────────────────────────
    inactive_users: list[dict] = []
    low_users: list[dict] = []

    for scope in scopes:
        seats_data = data_collector.load_latest("seats", scope)
        billing = data_collector.load_latest("billing", scope)
        price = billing.get("_detected_price_per_seat", 19.0) if billing else 19.0

        if not seats_data:
            continue

        # Build usage lookup from user_usage already aggregated
        usage_lookup = {u["login"]: u for u in user_usage if u["org"] == scope}

        for s in seats_data.get("seats", []):
            assignee = (s.get("assignee") or {})
            login = assignee.get("login", "")
            if not login:
                continue
            last = s.get("last_activity_at")
            teams_obj = s.get("assigning_team")
            team_name = (teams_obj.get("name", "") if isinstance(teams_obj, dict)
                         else (teams_obj[0].get("name", "") if isinstance(teams_obj, list) and teams_obj else ""))

            days_inactive = None
            if last:
                try:
                    last_dt = datetime.fromisoformat(last.replace("Z", "+00:00"))
                    days_inactive = (now - last_dt).days
                except (ValueError, TypeError):
                    pass

            u_info = usage_lookup.get(login, {})
            days_active_in_period = u_info.get("days_active", 0)

            # Never used in period
            if days_active_in_period == 0:
                inactive_users.append({
                    "login": login, "org": scope, "team": team_name,
                    "last_active": last[:10] if last else "Never",
                    "days_inactive": days_inactive if days_inactive is not None else 9999,
                    "cost_per_month": price,
                    "recommendation": "Revoke seat",
                })
            elif days_active_in_period <= 2:
                low_users.append({
                    "login": login, "org": scope, "team": team_name,
                    "days_active_in_period": days_active_in_period,
                    "last_active": last[:10] if last else "",
                    "cost_per_month": price,
                    "recommendation": "Send reminder / review",
                })

    inactive_users.sort(key=lambda x: -x["days_inactive"])
    low_users.sort(key=lambda x: x["days_active_in_period"])

    potential_savings = sum(u["cost_per_month"] for u in inactive_users)

    # Identify reallocation candidates: inactive users from orgs that also have low-usage active users
    # who might benefit from a reassignment (org has inactive seats available)
    reallocation_candidates: list[dict] = []
    inactive_logins_by_org: dict[str, set] = defaultdict(set)
    for u in inactive_users:
        inactive_logins_by_org[u["org"]].add(u["login"])

    for u in inactive_users:
        days_inactive = u["days_inactive"]
        action = "Thu hồi & cấp lại" if days_inactive >= 30 else "Thu hồi"
        reallocation_candidates.append({**u, "action_type": action})

    reallocation_candidates.sort(key=lambda x: -x["days_inactive"])
    total_reallocatable = sum(1 for r in reallocation_candidates if r["action_type"] == "Thu hồi & cấp lại")

    # ── Section 8: Premium Requests ───────────────────────────────────────────
    premium_model_map: dict[str, dict] = defaultdict(lambda: {
        "requests": 0, "gross_amount": 0.0, "net_amount": 0.0
    })
    total_premium_requests = 0
    total_premium_cost = 0.0
    for scope in scopes:
        pr_data = data_collector.load_latest("premium_requests", scope)
        if not pr_data:
            continue
        for rec in pr_data.get("usages", pr_data.get("records", [])):
            # handle both list-of-daily and flat formats
            if isinstance(rec, dict):
                month_str = (rec.get("month") or rec.get("day", ""))[:7]
                if not any(month_str == m for m in months):
                    continue
                for item in rec.get("breakdown", rec.get("items", [])):
                    model = item.get("model_name", item.get("model", "unknown"))
                    reqs = item.get("total_requests", item.get("requests", 0))
                    gross = item.get("gross_amount_usd", item.get("gross", 0.0))
                    net = item.get("net_amount_usd", item.get("net", 0.0))
                    premium_model_map[model]["requests"] += reqs
                    premium_model_map[model]["gross_amount"] += gross
                    premium_model_map[model]["net_amount"] += net
                    total_premium_requests += reqs
                    total_premium_cost += net

    premium_rows = sorted(
        [{"model": k, **v} for k, v in premium_model_map.items()],
        key=lambda x: -x["requests"]
    )

    # ── Section Quarterly: Month-by-Month Breakdown ────────────────────────────
    # When period_type == "quarterly", compute per-month summaries for Sections 3, 5, 6
    monthly_breakdown: list[dict] = []
    if period_type == "quarterly":
        import calendar as _cal
        for m_str in months:
            m_year, m_num = int(m_str[:4]), int(m_str[5:7])
            m_name = _cal.month_abbr[m_num]
            m_months = [m_str]  # single month filter

            m_interactions = 0
            m_code_gen = 0
            m_code_accept = 0
            m_dau_sum = 0
            m_days = 0
            m_chat = 0
            m_active_users: set[str] = set()

            for scope in scopes:
                usage = data_collector.load_latest("usage", scope)
                if usage:
                    for rec in usage.get("records", []):
                        for dt in rec.get("day_totals", []):
                            if not _in_period(dt.get("day", ""), m_months):
                                continue
                            m_interactions += dt.get("user_initiated_interaction_count", 0)
                            m_code_gen += dt.get("code_generation_activity_count", 0)
                            m_code_accept += dt.get("code_acceptance_activity_count", 0)
                            m_dau_sum += dt.get("daily_active_users", 0)
                            m_days += 1
                            for fb in dt.get("totals_by_feature", []):
                                fname = (fb.get("feature") or "").lower()
                                if "chat" in fname:
                                    m_chat += fb.get("user_initiated_interaction_count", 0)

                uu_data = data_collector.load_latest("usage_users", scope)
                if uu_data:
                    for rec in uu_data.get("records", []):
                        if _in_period(rec.get("day", ""), m_months):
                            login = rec.get("user_login", "")
                            if login:
                                m_active_users.add(login)

            monthly_breakdown.append({
                "month": m_str,
                "month_label": f"{m_name} {m_year}",
                "interactions": m_interactions,
                "code_suggestions": m_code_gen,
                "accepted_suggestions": m_code_accept,
                "acceptance_rate": round(m_code_accept / m_code_gen * 100, 1) if m_code_gen > 0 else 0.0,
                "avg_dau": round(m_dau_sum / m_days, 1) if m_days > 0 else 0,
                "active_users": len(m_active_users),
                "chat_interactions": m_chat,
            })

    return {
        "label": label,
        "period_type": period_type,
        "year": year,
        "period": period,
        "months": months,
        "generated_at": now.strftime("%Y-%m-%d %H:%M UTC"),
        # section 1
        "license": {
            "rows": license_rows,
            "team_rows": team_rows,
            "total_seats": total_seats_all,
            "total_active": total_active_all,
            "total_inactive": total_seats_all - total_active_all,
            "utilization_pct": round(total_active_all / total_seats_all * 100, 1) if total_seats_all > 0 else 0.0,
            "total_monthly_cost": round(total_cost_all, 2),
        },
        # section 2
        "user_usage": user_usage,
        # section 3
        "copilot_stats": {
            "total_interactions": total_interactions,
            "total_code_suggestions": total_code_gen,
            "total_accepted_suggestions": total_code_accept,
            "acceptance_rate": acceptance_rate,
            "total_loc_suggested": total_loc_suggested,
            "total_loc_accepted": total_loc_accepted,
            "avg_dau": avg_dau,
            "total_chat_interactions": total_chat_interactions,
            "peak_chat_users": peak_chat_users,
            "total_pr_summary_interactions": total_pr_summary_interactions,
            "chat_pct_of_total": round(total_chat_interactions / total_interactions * 100, 1) if total_interactions > 0 else 0.0,
        },
        # section 4
        "technical": {
            "models": model_rows,
            "ides": ide_rows,
            "languages": lang_rows,
            "features": feature_rows,
        },
        # section 5
        "org_stats": org_stats,
        "team_usage_stats": team_usage_rows,
        # section 6
        "trends": daily_trend,
        "weekly_trends": weekly_trend,
        # section 7
        "optimization": {
            "inactive_users": inactive_users,
            "low_usage_users": low_users,
            "reallocation_candidates": reallocation_candidates,
            "potential_monthly_savings": round(potential_savings, 2),
            "total_reallocatable": total_reallocatable,
        },
        # section 8
        "premium_requests": {
            "rows": premium_rows,
            "total_requests": total_premium_requests,
            "total_cost": round(total_premium_cost, 2),
        },
        # quarterly monthly breakdown
        "monthly_breakdown": monthly_breakdown,
    }


# ══════════════════════════════════════════════════════════════════════════════
# HTML RENDERER
# ══════════════════════════════════════════════════════════════════════════════

def _render_html(d: dict) -> str:
    label = d["label"]
    gen = d["generated_at"]
    lic = d["license"]
    cs = d["copilot_stats"]
    opt = d["optimization"]

    # ── Section 1: Quản lý License ────────────────────────────────────────────
    sec1_kpis = _kpi_row(
        _kpi("Tổng số License", _num(lic["total_seats"])),
        _kpi("License Active", _num(lic["total_active"]), _GREEN),
        _kpi("License Chưa dùng", _num(lic["total_inactive"]), _RED),
        _kpi("Tỷ lệ sử dụng", _pct(lic["utilization_pct"]),
             _GREEN if lic["utilization_pct"] >= 70 else _YELLOW),
        _kpi("Chi phí tháng", _money(lic["total_monthly_cost"]), _YELLOW),
    )
    sec1_table = _table(
        ["Org/Enterprise", "Gói", "Tổng License", "Active", "Chưa dùng", "Tỷ lệ sử dụng", "Chi phí/tháng"],
        [[r["org"], r["plan"], _num(r["total_seats"]), _num(r["active_seats"]),
          _num(r["inactive_seats"]), _pct(r["utilization_pct"]), _money(r["monthly_cost"])]
         for r in lic["rows"]],
    )
    sec1_team = _table(
        ["Org", "Team/Đơn vị", "Tổng", "Active", "Chưa dùng", "Tỷ lệ sử dụng"],
        [[r["org"], r["team"], _num(r["total"]), _num(r["active"]),
          _num(r["inactive"]), _pct(r["utilization_pct"])]
         for r in sorted(lic["team_rows"], key=lambda x: -x["active"])],
    )
    sec1 = _section("1. Thống kê quản lý License", sec1_kpis + sec1_table +
                    "<h4 style='margin:16px 0 8px;color:#e6edf3'>Theo Team / Đơn vị</h4>" + sec1_team,
                    "🔑", anchor="sec1")

    # ── Section 2: Thống kê người dùng ───────────────────────────────────────
    active_users = [u for u in d["user_usage"] if u["days_active"] > 0]
    sec2_kpis = _kpi_row(
        _kpi("Người dùng Active", _num(len(active_users))),
        _kpi("TB ngày hoạt động", str(round(
            sum(u["days_active"] for u in active_users) / len(active_users), 1
        ) if active_users else 0)),
        _kpi("TB Acceptance Rate", _pct(
            round(sum(u["acceptance_rate"] for u in active_users) / len(active_users), 1)
            if active_users else 0
        ), _GREEN),
    )
    sec2_table = _table(
        ["Người dùng", "Org", "Team/Đơn vị", "Ngày bắt đầu", "Hoạt động gần nhất",
         "Ngày HĐ", "Tuần HĐ", "Tần suất", "TB tương tác/ngày",
         "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance %"],
        [[u["login"], u["org"], u["team"], u["created_at"], u["last_active"],
          u["days_active"], u["weeks_active"],
          u.get("usage_frequency", "-"), u.get("avg_interactions_per_active_day", 0),
          _num(u["interactions"]),
          _num(u["code_suggestions"]), _num(u["accepted_suggestions"]),
          _pct(u["acceptance_rate"])]
         for u in active_users],
    )
    sec2 = _section("2. Thống kê người dùng", sec2_kpis + sec2_table, "👤", anchor="sec2")

    # ── Section 3: Mức độ sử dụng Copilot ─────────────────────────────────────
    sec3_kpis = _kpi_row(
        _kpi("Tổng tương tác", _num(cs["total_interactions"])),
        _kpi("Tổng gợi ý code", _num(cs["total_code_suggestions"])),
        _kpi("Gợi ý được chấp nhận", _num(cs["total_accepted_suggestions"]), _GREEN),
        _kpi("Tỷ lệ chấp nhận", _pct(cs["acceptance_rate"]),
             _GREEN if cs["acceptance_rate"] >= 25 else _YELLOW),
        _kpi("Dòng code được gợi ý", _num(cs["total_loc_suggested"])),
        _kpi("Dòng code được chấp nhận", _num(cs["total_loc_accepted"]), _GREEN),
        _kpi("Trung bình DAU", str(cs["avg_dau"])),
    )
    sec3_chat_kpis = _kpi_row(
        _kpi("Tổng Chat/Prompt", _num(cs["total_chat_interactions"]), _ACCENT),
        _kpi("Tỷ lệ Chat / Tổng tương tác", _pct(cs["chat_pct_of_total"]), _ACCENT),
        _kpi("Peak Monthly Chat Users", _num(cs["peak_chat_users"])),
        _kpi("PR Summary Interactions", _num(cs["total_pr_summary_interactions"])),
    )
    sec3 = _section("3. Thống kê mức độ sử dụng Copilot",
                    sec3_kpis +
                    "<h4 style='margin:16px 0 8px;color:#e6edf3'>📨 Copilot Chat & PR Summary</h4>" +
                    sec3_chat_kpis, "📊", anchor="sec3")

    # ── Section 4: Thống kê kỹ thuật ──────────────────────────────────────────
    tech = d["technical"]
    sec4_model = _table(
        ["Model AI", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance %"],
        [[r["model"], _num(r["interactions"]), _num(r["code_gen"]),
          _num(r["code_accept"]), _pct(r["acceptance_rate"])]
         for r in tech["models"]],
    )
    sec4_ide = _table(
        ["IDE", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance %"],
        [[r["ide"], _num(r["interactions"]), _num(r["code_gen"]),
          _num(r["code_accept"]), _pct(r["acceptance_rate"])]
         for r in tech["ides"]],
    )
    sec4_lang = _table(
        ["Ngôn ngữ", "Gợi ý", "Đã chấp nhận", "Dòng được gợi ý", "Dòng được chấp nhận", "Acceptance %"],
        [[r["language"], _num(r["code_gen"]), _num(r["code_accept"]),
          _num(r["loc_suggested"]), _num(r["loc_accepted"]), _pct(r["acceptance_rate"])]
         for r in tech["languages"][:50]],
    )
    sec4_feat = _table(
        ["Tính năng", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance %"],
        [[r["feature"], _num(r["interactions"]), _num(r["code_gen"]),
          _num(r["code_accept"]), _pct(r["acceptance_rate"])]
         for r in tech["features"]],
    )
    _repo_note = (
        f'<div style="background:{_CARD2};border-left:3px solid {_YELLOW};'
        f'padding:10px 14px;margin:12px 0;border-radius:4px;font-size:12px;color:{_MUTED}">'
        f'ℹ️ <strong style="color:{_TEXT}">Lưu ý — Thống kê theo Repository/Project:</strong> '
        f'GitHub Copilot Usage Metrics API hiện tại <strong>không cung cấp</strong> dữ liệu chi tiết '
        f'theo repository/project. Mục <strong>5. Thống kê theo Team/Nhóm</strong> bên dưới là proxy '
        f'gần nhất có thể thay thế. Dữ liệu repository-level có thể được bổ sung khi GitHub API hỗ trợ.'
        f'</div>'
    )
    sec4 = _section("4. Thống kê kỹ thuật",
        _repo_note +
        "<h4 style='margin:12px 0 8px;color:#e6edf3'>Theo Model AI</h4>" + sec4_model +
        "<h4 style='margin:16px 0 8px;color:#e6edf3'>Theo IDE</h4>" + sec4_ide +
        "<h4 style='margin:16px 0 8px;color:#e6edf3'>Theo Ngôn ngữ lập trình (Top 50)</h4>" + sec4_lang +
        "<h4 style='margin:16px 0 8px;color:#e6edf3'>Theo Loại tính năng</h4>" + sec4_feat,
        "⚙️", anchor="sec4")

    # ── Section 5: Thống kê theo đơn vị ──────────────────────────────────────
    sec5_table = _table(
        ["Hạng", "Org/Đơn vị", "Tổng License", "User Active", "Tỷ lệ sử dụng",
         "Tương tác", "Gợi ý", "Đã chấp nhận", "Acceptance %", "Chi phí/tháng"],
        [[r["rank"], r["org"], _num(r["total_seats"]), _num(r["active_users_period"]),
          _pct(r["utilization_pct"]), _num(r["interactions"]),
          _num(r["code_suggestions"]), _num(r["accepted_suggestions"]),
          _pct(r["acceptance_rate"]), _money(r["monthly_cost"])]
         for r in d["org_stats"]],
    )
    sec5_team_table = _table(
        ["Org", "Team/Nhóm", "Tổng License", "User Active", "Tỷ lệ sử dụng",
         "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance %"],
        [[r["org"], r["team"], _num(r["total_seats"]), _num(r["active_users"]),
          _pct(r["utilization_pct"]), _num(r["interactions"]),
          _num(r["code_suggestions"]), _num(r["accepted_suggestions"]),
          _pct(r["acceptance_rate"])]
         for r in d["team_usage_stats"]],
    )
    sec5 = _section("5. Thống kê và đánh giá theo đơn vị",
                    sec5_table +
                    "<h4 style='margin:16px 0 8px;color:#e6edf3'>Theo Team / Nhóm (Usage Metrics)</h4>" +
                    sec5_team_table, "🏢", anchor="sec5")

    # ── Section 6: Xu hướng sử dụng ──────────────────────────────────────────
    trends = d["trends"]
    weekly_trends = d.get("weekly_trends", [])
    chart_dau = _svg_line(trends, "day", "dau", "Daily Active Users (DAU)", "dau")
    chart_wau = _svg_line(trends, "day", "wau", "Weekly Active Users (WAU)", "wau")
    chart_mau = _svg_line(trends, "day", "mau", "Monthly Active Users (MAU)", "mau")
    chart_int = _svg_line(trends, "day", "interactions", "Tương tác hàng ngày", "int")

    def _growth_badge(v) -> str:
        if v is None:
            return "-"
        color = _GREEN if v >= 0 else _RED
        arrow = "▲" if v >= 0 else "▼"
        return f'<span style="color:{color}">{arrow} {abs(v)}%</span>'

    weekly_rows = []
    for r in weekly_trends:
        weekly_rows.append([r["week"], r["avg_dau"], _num(r["interactions"]),
                            _num(r["code_gen"]), _growth_badge(r["growth_pct"])])

    weekly_table_html = (
        f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse">'
        f'<thead><tr>'
        + "".join(
            f'<th style="padding:8px 12px;text-align:left;color:{_MUTED};font-size:12px;'
            f'font-weight:600;border-bottom:1px solid {_BORDER}">{h}</th>'
            for h in ["Tuần", "TB DAU", "Tổng tương tác", "Gợi ý code", "Tăng trưởng WoW"]
        )
        + f'</tr></thead><tbody>'
    )
    for i, row in enumerate(weekly_rows):
        bg = _CARD if i % 2 == 0 else _CARD2
        weekly_table_html += f'<tr style="background:{bg}">'
        for j, cell in enumerate(row):
            weekly_table_html += (
                f'<td style="padding:7px 12px;font-size:13px;border-bottom:1px solid {_BORDER}">'
                + (cell if j == 4 else _e(cell))  # col 4 is pre-rendered HTML badge
                + "</td>"
            )
        weekly_table_html += "</tr>"
    weekly_table_html += "</tbody></table></div>"

    trend_table = _table(
        ["Ngày", "DAU", "WAU", "MAU", "Tương tác", "Gợi ý code", "DAU WoW%", "Tương tác WoW%"],
        [[r["day"], _num(r["dau"]), _num(r["wau"]), _num(r["mau"]),
          _num(r["interactions"]), _num(r["code_gen"]),
          f"{r['dau_wow_pct']:+.1f}%" if r.get("dau_wow_pct") is not None else "-",
          f"{r['int_wow_pct']:+.1f}%" if r.get("int_wow_pct") is not None else "-"]
         for r in trends],
    )

    # Quarterly: add month-by-month comparison table in sec6
    quarterly_sec6_html = ""
    mb = d.get("monthly_breakdown", [])
    if mb:
        prev_int = None
        mb_rows = []
        for r in mb:
            mom = None
            if prev_int is not None and prev_int > 0:
                mom = round((r["interactions"] - prev_int) / prev_int * 100, 1)
            mb_rows.append([
                r["month_label"],
                _num(r["active_users"]),
                _num(r["interactions"]),
                _num(r["code_suggestions"]),
                _num(r["accepted_suggestions"]),
                _pct(r["acceptance_rate"]),
                str(r["avg_dau"]),
                _growth_badge(mom),
            ])
            prev_int = r["interactions"]
        quarterly_sec6_html = (
            "<h4 style='margin:16px 0 8px;color:#e6edf3'>📅 So sánh từng tháng trong quý (MoM)</h4>"
            + f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse">'
            + f'<thead><tr>'
            + "".join(
                f'<th style="padding:8px 12px;text-align:left;color:{_MUTED};font-size:12px;'
                f'font-weight:600;border-bottom:1px solid {_BORDER}">{h}</th>'
                for h in ["Tháng", "User Active", "Tương tác", "Gợi ý", "Đã chấp nhận",
                          "Acceptance %", "TB DAU", "Tăng trưởng MoM"]
            )
            + f'</tr></thead><tbody>'
            + "".join(
                f'<tr style="background:{_CARD if i % 2 == 0 else _CARD2}">'
                + "".join(
                    f'<td style="padding:7px 12px;font-size:13px;border-bottom:1px solid {_BORDER}">'
                    + (cell if j == 7 else _e(cell)) + "</td>"
                    for j, cell in enumerate(row)
                )
                + "</tr>"
                for i, row in enumerate(mb_rows)
            )
            + "</tbody></table></div>"
        )

    sec6 = _section("6. Thống kê xu hướng sử dụng",
        f"<h4 style='margin:0 0 8px;color:#e6edf3'>Daily Active Users (DAU)</h4>{chart_dau}"
        f"<h4 style='margin:16px 0 8px;color:#e6edf3'>Weekly Active Users (WAU)</h4>{chart_wau}"
        f"<h4 style='margin:16px 0 8px;color:#e6edf3'>Monthly Active Users (MAU)</h4>{chart_mau}"
        f"<h4 style='margin:16px 0 8px;color:#e6edf3'>Tương tác hàng ngày</h4>{chart_int}"
        f"{quarterly_sec6_html}"
        f"<h4 style='margin:16px 0 8px;color:#e6edf3'>Tổng hợp theo tuần (với tăng trưởng WoW)</h4>"
        f"{weekly_table_html}"
        f"<h4 style='margin:16px 0 8px;color:#e6edf3'>Dữ liệu hàng ngày</h4>{trend_table}",
        "📈", anchor="sec6")

    # ── Section 7: Tối ưu License ─────────────────────────────────────────────
    opt_kpis = _kpi_row(
        _kpi("User không hoạt động (0 ngày)", _num(len(opt["inactive_users"])), _RED),
        _kpi("User ít dùng (≤2 ngày)", _num(len(opt["low_usage_users"])), _YELLOW),
        _kpi("License có thể tái phân bổ", _num(opt.get("total_reallocatable", 0)), _ACCENT),
        _kpi("Tiết kiệm tiềm năng/tháng", _money(opt["potential_monthly_savings"]), _GREEN),
    )
    realloc_table = _table(
        ["Người dùng", "Org", "Team", "Hoạt động cuối", "Số ngày không HĐ",
         "Chi phí/tháng", "Loại đề xuất"],
        [[u["login"], u["org"], u["team"], u["last_active"],
          u["days_inactive"] if u["days_inactive"] != 9999 else "Chưa bao giờ",
          _money(u["cost_per_month"]),
          f'<span style="color:{_ACCENT if u.get("action_type","") == "Thu hồi & cấp lại" else _RED}">'
          f'{_e(u.get("action_type", "Thu hồi"))}</span>']
         for u in opt.get("reallocation_candidates", opt["inactive_users"])],
    )
    low_table = _table(
        ["Người dùng", "Org", "Team", "Ngày HĐ trong kỳ", "Hoạt động cuối", "Chi phí/tháng", "Đề xuất"],
        [[u["login"], u["org"], u["team"], u["days_active_in_period"],
          u["last_active"], _money(u["cost_per_month"]), "Gửi nhắc nhở / xem xét"]
         for u in opt["low_usage_users"]],
    )
    sec7 = _section("7. Thống kê phục vụ tối ưu License",
        opt_kpis +
        "<h4 style='margin:0 0 8px;color:#e6edf3'>Danh sách user không hoạt động — Phân loại đề xuất</h4>"
        + f'<p style="font-size:12px;color:{_MUTED};margin-bottom:8px">'
        f'🔴 <strong>Thu hồi</strong>: user inactive &lt; 30 ngày | '
        f'🔵 <strong>Thu hồi &amp; cấp lại</strong>: inactive ≥ 30 ngày — license nên được tái phân bổ</p>'
        + realloc_table
        + "<h4 style='margin:16px 0 8px;color:#e6edf3'>Danh sách user ít sử dụng (≤2 ngày)</h4>"
        + low_table,
        "🔧", anchor="sec7")

    # ── Section 8: Premium Requests ───────────────────────────────────────────
    pr_data = d["premium_requests"]
    if pr_data["total_requests"] > 0:
        pr_kpis = _kpi_row(
            _kpi("Tổng Premium Requests", _num(pr_data["total_requests"]), _YELLOW),
            _kpi("Tổng chi phí Premium ($)", _money(pr_data["total_cost"]), _RED),
        )
        pr_table = _table(
            ["Model AI", "Số Requests", "Chi phí gộc ($)", "Chi phí thực ($)"],
            [[r["model"], _num(r["requests"]), _money(r["gross_amount"]), _money(r["net_amount"])]
             for r in pr_data["rows"]],
        )
        sec8_content = pr_kpis + pr_table
    else:
        sec8_content = (f'<p style="color:{_MUTED};padding:16px 0">Chưa có dữ liệu Premium Requests '
                        f'cho kỳ này. Vui lòng tải lên CSV hoặc đồng bộ từ GitHub API.</p>')
    sec8 = _section("8. Thống kê Premium Requests", sec8_content, "💎", anchor="sec8")

    # ── sticky Table of Contents ───────────────────────────────────────────────
    _toc = (
        f'<nav style="position:sticky;top:0;z-index:100;background:{_BG};'
        f'border-bottom:1px solid {_BORDER};padding:8px 0 8px 0;margin-bottom:16px;'
        f'display:flex;flex-wrap:wrap;gap:4px 12px;font-size:12px">'
        + "".join(
            f'<a href="#{anchor}" style="color:{_MUTED};text-decoration:none;white-space:nowrap">'
            f'{icon} {title}</a>'
            for anchor, icon, title in [
                ("sec1", "🔑", "1. License"),
                ("sec2", "👤", "2. Người dùng"),
                ("sec3", "📊", "3. Mức độ dùng"),
                ("sec4", "⚙️", "4. Kỹ thuật"),
                ("sec5", "🏢", "5. Theo đơn vị"),
                ("sec6", "📈", "6. Xu hướng"),
                ("sec7", "🔧", "7. Tối ưu"),
                ("sec8", "💎", "8. Premium"),
            ]
        )
        + '</nav>'
    )

    # ── assemble full page ────────────────────────────────────────────────────
    body = (f'<h1>Báo cáo định kỳ: {_e(label)}</h1>'
            f'<p class="subtitle">Ngày tạo: {_e(gen)} | OctoFinance AI FinOps</p>'
            f'{_toc}'
            f'{sec1}{sec2}{sec3}{sec4}{sec5}{sec6}{sec7}{sec8}')

    return (f'<!DOCTYPE html><html lang="vi"><head>'
            f'<meta charset="UTF-8">'
            f'<meta name="viewport" content="width=device-width,initial-scale=1">'
            f'<title>Báo cáo định kỳ {_e(label)}</title>'
            f'<style>{_CSS}</style>'
            f'</head><body>{body}</body></html>')


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL RENDERER
# ══════════════════════════════════════════════════════════════════════════════

def _render_xlsx(d: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # colour scheme
    HDR_FILL = PatternFill("solid", fgColor="0D1117")
    HDR_FONT = Font(bold=True, color="539BF5", size=10)
    HDR_BORDER = Border(
        bottom=Side(style="thin", color="30363D"),
    )
    ROW_FONT = Font(color="E6EDF3", size=10)
    EVEN_FILL = PatternFill("solid", fgColor="161B22")
    ODD_FILL  = PatternFill("solid", fgColor="21262D")
    TITLE_FONT = Font(bold=True, color="E6EDF3", size=12)

    def _add_sheet(name: str, headers: list[str], rows: list[list]) -> None:
        ws = wb.create_sheet(title=name[:31])

        # title row
        ws.append([name])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        ws["A1"].font = TITLE_FONT
        ws["A1"].fill = PatternFill("solid", fgColor="0D1117")
        ws["A1"].alignment = Alignment(horizontal="left")
        ws.append([])  # blank

        # header row
        ws.append(headers)
        hdr_row = ws.max_row
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=hdr_row, column=col)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.border = HDR_BORDER
            cell.alignment = Alignment(horizontal="left")

        # data rows
        for i, row in enumerate(rows):
            ws.append(row)
            r = ws.max_row
            fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
            for col in range(1, len(row) + 1):
                cell = ws.cell(row=r, column=col)
                cell.font = ROW_FONT
                cell.fill = fill
                cell.alignment = Alignment(horizontal="left")

        # auto-width (cap at 50)
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(str(headers[col_idx - 1]))
            for row in rows:
                if col_idx <= len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1] or "")))
            ws.column_dimensions[letter].width = min(max_len + 2, 50)

        # freeze header rows
        ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    lic = d["license"]
    cs = d["copilot_stats"]
    tech = d["technical"]
    opt = d["optimization"]

    # Sheet 1: Thống kê quản lý License
    _add_sheet("1_Quan_Ly_License",
        ["Org/Enterprise", "Gói", "Tổng License", "License Active", "Chưa sử dụng",
         "Tỷ lệ sử dụng %", "Chi phí/tháng ($)"],
        [[r["org"], r["plan"], r["total_seats"], r["active_seats"], r["inactive_seats"],
          r["utilization_pct"], r["monthly_cost"]]
         for r in lic["rows"]]
    )

    # Sheet 1b: License theo Team
    _add_sheet("1b_License_Theo_Team",
        ["Org", "Team/Đơn vị", "Tổng License", "Active", "Chưa dùng", "Tỷ lệ sử dụng %"],
        [[r["org"], r["team"], r["total"], r["active"], r["inactive"], r["utilization_pct"]]
         for r in sorted(lic["team_rows"], key=lambda x: -x["active"])]
    )

    # Sheet 2: Thống kê người dùng
    _add_sheet("2_Thong_Ke_Nguoi_Dung",
        ["Người dùng", "Org", "Team/Đơn vị", "Ngày bắt đầu dùng", "Hoạt động gần nhất",
         "Ngày hoạt động", "Tuần hoạt động", "Tần suất", "TB tương tác/ngày",
         "Tổng tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[u["login"], u["org"], u["team"], u["created_at"], u["last_active"],
          u["days_active"], u["weeks_active"],
          u.get("usage_frequency", "-"), u.get("avg_interactions_per_active_day", 0),
          u["interactions"],
          u["code_suggestions"], u["accepted_suggestions"], u["acceptance_rate"]]
         for u in d["user_usage"]]
    )

    # Sheet 3: Mức độ sử dụng Copilot (KPI tổng hợp)
    _add_sheet("3_Muc_Do_Su_Dung_Copilot",
        ["Chỉ số", "Giá trị"],
        [
            ["Tổng tương tác người dùng", cs["total_interactions"]],
            ["Tổng gợi ý code", cs["total_code_suggestions"]],
            ["Tổng gợi ý được chấp nhận", cs["total_accepted_suggestions"]],
            ["Tỷ lệ chấp nhận (%)", cs["acceptance_rate"]],
            ["Tổng dòng code được gợi ý", cs["total_loc_suggested"]],
            ["Tổng dòng code được chấp nhận", cs["total_loc_accepted"]],
            ["Trung bình Daily Active Users", cs["avg_dau"]],
            ["--- Copilot Chat ---", ""],
            ["Tổng Chat/Prompt Interactions", cs["total_chat_interactions"]],
            ["Tỷ lệ Chat / Tổng tương tác (%)", cs["chat_pct_of_total"]],
            ["Peak Monthly Chat Users", cs["peak_chat_users"]],
            ["PR Summary Interactions", cs["total_pr_summary_interactions"]],
        ]
    )

    # Sheet 4a: Theo Model AI
    _add_sheet("4a_Theo_Model_AI",
        ["Model AI", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[r["model"], r["interactions"], r["code_gen"], r["code_accept"], r["acceptance_rate"]]
         for r in tech["models"]]
    )

    # Sheet 4b: Theo IDE
    _add_sheet("4b_Theo_IDE",
        ["IDE", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[r["ide"], r["interactions"], r["code_gen"], r["code_accept"], r["acceptance_rate"]]
         for r in tech["ides"]]
    )

    # Sheet 4c: Theo Ngôn ngữ lập trình
    _add_sheet("4c_Theo_Ngon_Ngu",
        ["Ngôn ngữ", "Gợi ý", "Đã chấp nhận", "Dòng được gợi ý", "Dòng được chấp nhận", "Acceptance Rate %"],
        [[r["language"], r["code_gen"], r["code_accept"],
          r["loc_suggested"], r["loc_accepted"], r["acceptance_rate"]]
         for r in tech["languages"]]
    )

    # Sheet 4d: Theo Tính năng
    _add_sheet("4d_Theo_Tinh_Nang",
        ["Tính năng", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[r["feature"], r["interactions"], r["code_gen"], r["code_accept"], r["acceptance_rate"]]
         for r in tech["features"]]
    )

    # Sheet 5: Thống kê theo đơn vị (với Rank)
    _add_sheet("5_Thong_Ke_Theo_Don_Vi",
        ["Hạng", "Org/Đơn vị", "Tổng License", "User Active (kỳ)", "Tỷ lệ sử dụng %",
         "Tổng tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %", "Chi phí/tháng ($)"],
        [[r["rank"], r["org"], r["total_seats"], r["active_users_period"], r["utilization_pct"],
          r["interactions"], r["code_suggestions"], r["accepted_suggestions"],
          r["acceptance_rate"], r["monthly_cost"]]
         for r in d["org_stats"]]
    )

    # Sheet 5b: Thống kê theo Team (Usage Metrics)
    _add_sheet("5b_Thong_Ke_Theo_Team",
        ["Org", "Team/Nhóm", "Tổng License", "User Active", "Tỷ lệ sử dụng %",
         "Tổng tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[r["org"], r["team"], r["total_seats"], r["active_users"], r["utilization_pct"],
          r["interactions"], r["code_suggestions"], r["accepted_suggestions"], r["acceptance_rate"]]
         for r in d["team_usage_stats"]]
    )

    # Sheet 6a: Xu hướng hàng ngày
    _add_sheet("6a_Xu_Huong_Hang_Ngay",
        ["Ngày", "DAU", "WAU", "MAU", "Tổng tương tác", "Gợi ý code",
         "DAU WoW %", "Tương tác WoW %"],
        [[r["day"], r["dau"], r["wau"], r["mau"], r["interactions"], r["code_gen"],
          r["dau_wow_pct"] if r.get("dau_wow_pct") is not None else "",
          r["int_wow_pct"] if r.get("int_wow_pct") is not None else ""]
         for r in d["trends"]]
    )

    # Sheet 6b: Xu hướng hàng tuần (với tăng trưởng WoW)
    _add_sheet("6b_Xu_Huong_Hang_Tuan",
        ["Tuần", "TB DAU", "Tổng tương tác", "Gợi ý code", "Tăng trưởng WoW %"],
        [[r["week"], r["avg_dau"], r["interactions"], r["code_gen"],
          r["growth_pct"] if r["growth_pct"] is not None else ""]
         for r in d.get("weekly_trends", [])]
    )

    # Sheet 6c: Monthly breakdown (quarterly reports only)
    mb = d.get("monthly_breakdown", [])
    if mb:
        prev_int = None
        mb_rows = []
        for r in mb:
            mom = None
            if prev_int is not None and prev_int > 0:
                mom = round((r["interactions"] - prev_int) / prev_int * 100, 1)
            mb_rows.append([
                r["month_label"], r["active_users"], r["interactions"],
                r["code_suggestions"], r["accepted_suggestions"],
                r["acceptance_rate"], r["avg_dau"], mom if mom is not None else "",
            ])
            prev_int = r["interactions"]
        _add_sheet("6c_Theo_Thang_Trong_Quy",
            ["Tháng", "User Active", "Tổng tương tác", "Gợi ý code", "Đã chấp nhận",
             "Acceptance Rate %", "TB DAU", "Tăng trưởng MoM %"],
            mb_rows
        )

    # Sheet 7a: User không hoạt động + phân loại đề xuất
    realloc = opt.get("reallocation_candidates", opt["inactive_users"])
    _add_sheet("7a_Phan_Loai_De_Xuat",
        ["Người dùng", "Org", "Team", "Hoạt động cuối", "Số ngày không HĐ",
         "Chi phí/tháng ($)", "Loại đề xuất"],
        [[u["login"], u["org"], u["team"], u["last_active"],
          u["days_inactive"] if u["days_inactive"] != 9999 else "Chưa bao giờ",
          u["cost_per_month"], u.get("action_type", "Thu hồi")]
         for u in realloc]
    )

    # Sheet 7b: User ít sử dụng
    _add_sheet("7b_User_It_Su_Dung",
        ["Người dùng", "Org", "Team", "Ngày HĐ trong kỳ",
         "Hoạt động cuối", "Chi phí/tháng ($)", "Đề xuất"],
        [[u["login"], u["org"], u["team"], u["days_active_in_period"],
          u["last_active"], u["cost_per_month"], "Gửi nhắc nhở / xem xét"]
         for u in opt["low_usage_users"]]
    )

    # Sheet 8: Premium Requests
    pr = d["premium_requests"]
    _add_sheet("8_Premium_Requests",
        ["Chỉ số", "Giá trị"],
        [
            ["Tổng Premium Requests", pr["total_requests"]],
            ["Tổng chi phí Premium ($)", pr["total_cost"]],
        ]
    )
    if pr["rows"]:
        _add_sheet("8b_Premium_Theo_Model",
            ["Model AI", "Số Requests", "Chi phí gộc ($)", "Chi phí thực ($)"],
            [[r["model"], r["requests"], round(r["gross_amount"], 4), round(r["net_amount"], 4)]
             for r in pr["rows"]]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def _render_csv(d: dict) -> bytes:
    """Render all report sections into a single multi-section CSV (UTF-8 with BOM for Excel)."""
    import csv as _csv

    buf = io.StringIO()
    w = _csv.writer(buf)

    def _section_header(title: str) -> None:
        w.writerow([])
        w.writerow([f"=== {title} ==="])

    def _table(headers: list[str], rows: list[list]) -> None:
        w.writerow(headers)
        for row in rows:
            w.writerow(row)

    lic  = d["license"]
    cs   = d["copilot_stats"]
    tech = d["technical"]
    opt  = d["optimization"]
    pr   = d["premium_requests"]
    label = d["label"]

    # Title
    w.writerow([f"Báo cáo định kỳ GitHub Copilot — {label}"])
    w.writerow([f"Tạo lúc: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])

    # 1. License
    _section_header("1. Thống kê quản lý License")
    _table(
        ["Org/Enterprise", "Gói", "Tổng License", "License Active", "Chưa sử dụng",
         "Tỷ lệ sử dụng %", "Chi phí/tháng ($)"],
        [[r["org"], r["plan"], r["total_seats"], r["active_seats"], r["inactive_seats"],
          r["utilization_pct"], r["monthly_cost"]] for r in lic["rows"]]
    )
    w.writerow([])
    w.writerow(["License theo Team"])
    _table(
        ["Org", "Team/Đơn vị", "Tổng License", "Active", "Chưa dùng", "Tỷ lệ sử dụng %"],
        [[r["org"], r["team"], r["total"], r["active"], r["inactive"], r["utilization_pct"]]
         for r in sorted(lic["team_rows"], key=lambda x: -x["active"])]
    )

    # 2. Người dùng
    _section_header("2. Thống kê người dùng")
    _table(
        ["Người dùng", "Org", "Team/Đơn vị", "Ngày bắt đầu dùng", "Hoạt động gần nhất",
         "Ngày hoạt động", "Tuần hoạt động", "Tần suất", "TB tương tác/ngày",
         "Tổng tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[u["login"], u["org"], u["team"], u["created_at"], u["last_active"],
          u["days_active"], u["weeks_active"],
          u.get("usage_frequency", "-"), u.get("avg_interactions_per_active_day", 0),
          u["interactions"], u["code_suggestions"], u["accepted_suggestions"],
          u["acceptance_rate"]] for u in d["user_usage"]]
    )

    # 3. Mức độ sử dụng
    _section_header("3. Thống kê mức độ sử dụng Copilot")
    _table(
        ["Chỉ số", "Giá trị"],
        [
            ["Tổng tương tác người dùng", cs["total_interactions"]],
            ["Tổng gợi ý code", cs["total_code_suggestions"]],
            ["Tổng gợi ý được chấp nhận", cs["total_accepted_suggestions"]],
            ["Tỷ lệ chấp nhận (%)", cs["acceptance_rate"]],
            ["Tổng dòng code được gợi ý", cs["total_loc_suggested"]],
            ["Tổng dòng code được chấp nhận", cs["total_loc_accepted"]],
            ["Trung bình Daily Active Users", cs["avg_dau"]],
            ["Tổng Chat/Prompt Interactions", cs["total_chat_interactions"]],
            ["Tỷ lệ Chat / Tổng tương tác (%)", cs["chat_pct_of_total"]],
            ["Peak Monthly Chat Users", cs["peak_chat_users"]],
            ["PR Summary Interactions", cs["total_pr_summary_interactions"]],
        ]
    )

    # 4. Kỹ thuật
    _section_header("4a. Theo Model AI")
    _table(
        ["Model AI", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[r["model"], r["interactions"], r["code_gen"], r["code_accept"], r["acceptance_rate"]]
         for r in tech["models"]]
    )

    _section_header("4b. Theo IDE")
    _table(
        ["IDE", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[r["ide"], r["interactions"], r["code_gen"], r["code_accept"], r["acceptance_rate"]]
         for r in tech["ides"]]
    )

    _section_header("4c. Theo Ngôn ngữ lập trình")
    _table(
        ["Ngôn ngữ", "Gợi ý", "Đã chấp nhận", "Dòng được gợi ý", "Dòng được chấp nhận", "Acceptance Rate %"],
        [[r["language"], r["code_gen"], r["code_accept"],
          r["loc_suggested"], r["loc_accepted"], r["acceptance_rate"]] for r in tech["languages"]]
    )

    _section_header("4d. Theo Tính năng")
    _table(
        ["Tính năng", "Tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[r["feature"], r["interactions"], r["code_gen"], r["code_accept"], r["acceptance_rate"]]
         for r in tech["features"]]
    )
    w.writerow(["[Lưu ý] GitHub Copilot Usage API không cung cấp dữ liệu theo Repository/Project."
                " Vui lòng tham khảo mục 5 (Theo Team) làm proxy thay thế."])

    # 5. Theo đơn vị
    _section_header("5. Thống kê và đánh giá theo đơn vị")
    _table(
        ["Hạng", "Org/Đơn vị", "Tổng License", "User Active (kỳ)", "Tỷ lệ sử dụng %",
         "Tổng tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %", "Chi phí/tháng ($)"],
        [[r["rank"], r["org"], r["total_seats"], r["active_users_period"], r["utilization_pct"],
          r["interactions"], r["code_suggestions"], r["accepted_suggestions"],
          r["acceptance_rate"], r["monthly_cost"]] for r in d["org_stats"]]
    )

    _section_header("5b. Thống kê theo Team")
    _table(
        ["Org", "Team/Nhóm", "Tổng License", "User Active", "Tỷ lệ sử dụng %",
         "Tổng tương tác", "Gợi ý code", "Đã chấp nhận", "Acceptance Rate %"],
        [[r["org"], r["team"], r["total_seats"], r["active_users"], r["utilization_pct"],
          r["interactions"], r["code_suggestions"], r["accepted_suggestions"], r["acceptance_rate"]]
         for r in d["team_usage_stats"]]
    )

    # 6. Xu hướng
    _section_header("6a. Xu hướng hàng ngày")
    _table(
        ["Ngày", "DAU", "WAU", "MAU", "Tổng tương tác", "Gợi ý code",
         "DAU WoW %", "Tương tác WoW %"],
        [[r["day"], r["dau"], r["wau"], r["mau"], r["interactions"], r["code_gen"],
          r["dau_wow_pct"] if r.get("dau_wow_pct") is not None else "",
          r["int_wow_pct"] if r.get("int_wow_pct") is not None else ""]
         for r in d["trends"]]
    )

    _section_header("6b. Xu hướng hàng tuần")
    _table(
        ["Tuần", "TB DAU", "Tổng tương tác", "Gợi ý code", "Tăng trưởng WoW %"],
        [[r["week"], r["avg_dau"], r["interactions"], r["code_gen"],
          r["growth_pct"] if r["growth_pct"] is not None else ""]
         for r in d.get("weekly_trends", [])]
    )

    # 6c. Quarterly monthly breakdown
    mb = d.get("monthly_breakdown", [])
    if mb:
        _section_header("6c. So sánh từng tháng trong quý (MoM)")
        prev_int = None
        mb_rows = []
        for r in mb:
            mom = None
            if prev_int is not None and prev_int > 0:
                mom = round((r["interactions"] - prev_int) / prev_int * 100, 1)
            mb_rows.append([
                r["month_label"], r["active_users"], r["interactions"],
                r["code_suggestions"], r["accepted_suggestions"],
                r["acceptance_rate"], r["avg_dau"], mom if mom is not None else "",
            ])
            prev_int = r["interactions"]
        _table(
            ["Tháng", "User Active", "Tổng tương tác", "Gợi ý code", "Đã chấp nhận",
             "Acceptance Rate %", "TB DAU", "Tăng trưởng MoM %"],
            mb_rows
        )

    # 7. Tối ưu
    _section_header("7. Tối ưu License — Phân loại đề xuất")
    realloc = opt.get("reallocation_candidates", opt["inactive_users"])
    _table(
        ["Người dùng", "Org", "Team", "Hoạt động cuối", "Số ngày không HĐ",
         "Chi phí/tháng ($)", "Loại đề xuất"],
        [[u["login"], u["org"], u["team"], u["last_active"],
          u["days_inactive"] if u["days_inactive"] != 9999 else "Chưa bao giờ",
          u["cost_per_month"], u.get("action_type", "Thu hồi")] for u in realloc]
    )

    _section_header("7b. User ít sử dụng")
    _table(
        ["Người dùng", "Org", "Team", "Ngày HĐ trong kỳ",
         "Hoạt động cuối", "Chi phí/tháng ($)", "Đề xuất"],
        [[u["login"], u["org"], u["team"], u["days_active_in_period"],
          u["last_active"], u["cost_per_month"], "Gửi nhắc nhở / xem xét"]
         for u in opt["low_usage_users"]]
    )

    # 8. Premium Requests
    _section_header("8. Premium Requests")
    _table(
        ["Chỉ số", "Giá trị"],
        [
            ["Tổng Premium Requests", pr["total_requests"]],
            ["Tổng chi phí Premium ($)", pr["total_cost"]],
        ]
    )
    if pr["rows"]:
        w.writerow([])
        w.writerow(["Premium theo Model AI"])
        _table(
            ["Model AI", "Số Requests", "Chi phí gộc ($)", "Chi phí thực ($)"],
            [[r["model"], r["requests"], round(r["gross_amount"], 4), round(r["net_amount"], 4)]
             for r in pr["rows"]]
        )

    # Return UTF-8 with BOM so Excel opens Vietnamese correctly
    return ("\ufeff" + buf.getvalue()).encode("utf-8")

def generate_periodic_report(
    data_collector,
    period_type: str,
    year: int,
    period: int,
    fmt: str = "html",
    org_filter: list[str] | None = None,
    all_scope_names: list[str] | None = None,
) -> tuple[bytes, str, str]:
    """Generate a single-file periodic report.

    Args:
        fmt: Output format — 'html', 'csv', or 'xlsx'.

    Returns:
        Tuple of (file_bytes, filename, media_type).
    """
    org_filter = org_filter or []
    data = _aggregate(data_collector, period_type, year, period, org_filter, all_scope_names)
    label = data["label"]

    if fmt == "csv":
        file_bytes = _render_csv(data)
        filename = f"periodic-report-{label}.csv"
        media_type = "text/csv; charset=utf-8"
    elif fmt == "xlsx":
        file_bytes = _render_xlsx(data)
        filename = f"periodic-report-{label}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:  # html (default)
        file_bytes = _render_html(data).encode("utf-8")
        filename = f"periodic-report-{label}.html"
        media_type = "text/html; charset=utf-8"

    return file_bytes, filename, media_type


def generate_periodic_report_zip(
    data_collector,
    period_type: str,
    year: int,
    period: int,
    org_filter: list[str] | None = None,
    all_scope_names: list[str] | None = None,
) -> bytes:
    """Generate a ZIP containing HTML + XLSX periodic report (kept for backwards compat)."""
    org_filter = org_filter or []
    data = _aggregate(data_collector, period_type, year, period, org_filter, all_scope_names)
    label = data["label"]

    html_bytes = _render_html(data).encode("utf-8")
    xlsx_bytes = _render_xlsx(data)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"report-{label}.html", html_bytes)
        zf.writestr(f"report-{label}.xlsx", xlsx_bytes)
    return buf.getvalue()
